import re
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# PDF-path
pdf_path = "invoice.pdf"

# Pdf öffnen und die Seiten lesen
with pdfplumber.open(pdf_path) as pdf:
    pages = [page.extract_text() for page in pdf.pages]

# Die gesamte Information für file1
first_page_text = pages[0]  # Die erste Seite hat die Info für file1
lines = first_page_text.split("\n")
general_info = {
    "Invoice Number": next(line.split(":")[1].strip() for line in lines if "Nummer:" in line),
    "Customer Number": next(line.split(":")[1].strip() for line in lines if "Kunden-Nr.:" in line),
    "Invoice Date": next(line.split(":")[1].strip() for line in lines if "Datum:" in line),
    "Abrechnung bis": next(line.split(":")[1].split("Nettobetrag")[0].strip() for line in lines if "Abrechnung bis:" in line),
    "Anzahl Fahrz.": next(line.split("Anzahl Fahrz.")[1].split("MwSt")[0].strip() for line in lines if "Anzahl Fahrz." in line),
    "Nettobetrag": next(line.split("CHF")[1].strip() for line in lines if "Nettobetrag" in line),
    "MwSt %": next(line.split("MwSt")[1].split("%")[0].strip() for line in lines if "MwSt" in line),
    "MwSt CHF": next(line.split("CHF")[1].strip() for line in lines if "MwSt" in line),
    "Bruttobetrag": next(line.split("CHF")[1].strip() for line in lines if "Bruttobetrag" in line),
    "Zahlbar innerh.": next(line.replace("Zahlbar innerh.", "").replace("netto", "").strip() 
                              for line in lines if "Zahlbar innerh." in line)
}
df1 = pd.DataFrame([general_info])

# einfach speichern mit der gleichen Breite für alle Spalten 
with pd.ExcelWriter("file1.xlsx", engine="openpyxl") as writer:
    df1.to_excel(writer, index=False)
    worksheet = writer.sheets["Sheet1"]
    uniform_width = 25  # Breite für Spalten
    for col_num in range(1, len(df1.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = uniform_width

# Die Info für file2 wird hier herausgenommen
cars_data = []
for page_text in pages[1:]:  # Analyse von Seiten, gestartet mit der zweiten
    lines = page_text.split("\n")
    
    for i, line in enumerate(lines):
        # Raw mit Fahrzeugsinfo festlegen, entsprechend den Ziffern
        if any(keyword in line for keyword in ["BM", "VW", "DC", "FI", "NI", "ME", "VF", "VX", "GF", "PS", "EF", "TX", "AR", "RN", "WBA", "TMB", "VSS"]):
            line_parts = line.split()
            if len(line_parts) < 5:
                continue  # When the raw is too short

            ziffer = line_parts[0]
            if ziffer.strip() not in ["BM", "VW", "DC", "FI", "NI", "ME", "VF", "VX", "GF", "PS", "EF", "TX", "AR", "RN", "WBA", "TMB", "VSS"]:
                continue

            ziffer = line_parts[0]
            vin = line_parts[1]
            date = line_parts[2]
            model = " ".join(line_parts[3:-2])
            total = line_parts[-1].replace("CHF", "").strip()  # Summe
            invoice_nr = lines[i + 1].split()[0] if (i + 1) < len(lines) else ""
            
            # Raw herausnehmen (mit dem Faktor)
            location_line = ""
            for j in range(i+1, len(lines)):
                candidate = lines[j]
                if candidate.startswith("CH") and "Faktor" in candidate and "Ansatz" in candidate:
                    location_line = candidate
                    break
                    
            if location_line:
                pattern = r'CH\s+\S+\s+(.*?)\s+CH\s+\S+\s+(.*?)\s+Faktor\s+(\S+)\s+Ansatz\s+(\S+)'
                match = re.search(pattern, location_line)
                if match:
                    loadingcity = match.group(1)
                    delivercity = match.group(2)
                    faktor = match.group(3)
                    ansatz = match.group(4)
                else:
                    loadingcity = delivercity = faktor = ansatz = ""
            else:
                loadingcity = delivercity = faktor = ansatz = ""
            
            # LEERFAHRT kalkulieren
            leerfahrt = ""
            for j in range(i+1, min(i+6, len(lines))):
                if any(keyword in lines[j] for keyword in ["BM", "VW", "DC", "FI", "NI", "ME", "VF", "VX", "GF", "PS", "EF", "TX", "AR", "RN", "WBA", "TMB", "VSS"]):
                    break  # wenn neue dann abbrechen
                if "/ LEERFAHRT" in lines[j]:
                    leerfahrt = "OK"
                    break
            
            # Car Auktion Protokoll kalkulieren
            auktion_protokoll = ""
            for j in range(i+1, min(i+6, len(lines))):
                if "Car Auktion Protokoll" in lines[j]:
                    m = re.search(r"Car Auktion Protokoll(?:\s+CHF)?\s+([\d,]+)", lines[j])
                    if m:
                        auktion_protokoll = m.group(1)
                    break
            
            # Terminverein. Absender CarAukt kalkulieren
            terminverein = ""
            for j in range(i+1, min(i+6, len(lines))):
                if "Terminverein. Absender CarAukt" in lines[j]:
                    m = re.search(r"Terminverein\. Absender CarAukt(?:\s+CHF)?\s+([\d,]+)", lines[j])
                    if m:
                        terminverein = m.group(1)
                    break
            
            # Seilwinde-Zuschlag kalkulieren
            seilwinde_zuschlag = ""
            for j in range(i+1, min(i+6, len(lines))):
                if "Seilwinde-Zuschlag" in lines[j]:
                    m = re.search(r"Seilwinde-Zuschlag(?:\s+CHF)?\s+([\d,]+)", lines[j])
                    if m:
                        seilwinde_zuschlag = m.group(1)
                    break
            # Terminzuschlag kalkulieren
            terminzuschlag = ""
            for j in range(i+1, min(i+6, len(lines))):
                if "Terminzuschlag" in lines[j]:
                    m = re.search(r"Terminzuschlag(?:\s+CHF)?\s+([\d,]+)", lines[j])
                    if m:
                        terminzuschlag = m.group(1)
                    break
            
            # E-Übernahme kalkulieren
            EFahrzeug = ""
            for j in range(i+1, min(i+6, len(lines))):
                if "E-Fahrzeug" in lines[j]:
                    m = re.search(r"E-Fahrzeug(?:\s+CHF)?\s+([\d,]+)", lines[j])
                    if m:
                        EFahrzeug = m.group(1)
                    break

            # Auftraggeber kalkulieren
            inv_len = len(invoice_nr)
            if inv_len == 6:
                auftraggeber = "CA3"
            elif inv_len == 5:
                auftraggeber = "RRM"
            else:
                auftraggeber = "Fehler"

            cars_data.append({
                "Ziffer": ziffer,
                "VIN": vin,
                "Date": date,
                "Model": model,
                "Total": total,
                "InvoiceNr": invoice_nr,
                "Loadingcity": loadingcity,
                "Delivercity": delivercity,
                "Faktor": faktor,
                "Ansatz": ansatz,
                "LEERFAHRT": leerfahrt,
                "Car Auktion Protokoll": auktion_protokoll,
                "Terminverein. Absender CarAukt": terminverein,
                "Seilwinde": seilwinde_zuschlag,
                "Terminzuschlag": terminzuschlag,
                "E-Fahrzeug": EFahrzeug,
                "Auftraggeber": auftraggeber,
            })
            
df2 = pd.DataFrame(cars_data)

# Einfach speichern in die file2 Datei 
with pd.ExcelWriter("file2.xlsx", engine="openpyxl") as writer:
    df2.to_excel(writer, index=False)
    worksheet = writer.sheets["Sheet1"]
    uniform_width = 25
    for col_num in range(1, len(df2.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = uniform_width

# --------------------------------------------------
# Erstellung von file3.xlsx: Überprüfung und Aggregierung
#
# 1) "Anzahl Fahrz. Check":
#    Vergleicht den Wert aus file1 ("Anzahl Fahrz.") mit der
#    Anzahl der Datensätze in file2 (Anzahl der Einträge in cars_data).
#
# 2) "Nettobetrag Check":
#    Vergleicht den Nettobetrag aus file1 mit der aggregierten Summe,
#    die sich aus den Spalten ergibt:
#         Total,
#         LEERFAHRT,
#         Car Auktion Protokoll,
#         Terminverein. Absender CarAukt,
#         Seilwinde,
#         Terminzuschlag,
#         E-Fahrzeug.
#
# 3) "Summe Total+":
#    Zeigt die berechnete aggregierte Summe der oben genannten Spalten.
# --------------------------------------------------

def convert_german_number(value):
    """
    Konvertiert eine Zahl im deutschen Format in einen Float.
    Beispiel: "85.960,50" -> 85960.50, "850,60" -> 850.60
    """
    value = value.replace(".", "").replace(",", ".").replace(" ", "")
    return float(value)

# --------------------------------------------------
# 1. Extraktion "Summe ohne MwSt" aus der letzten PDF-Seite:
last_page_text = pages[-1]
summe_ohne_mwst = ""
for line in last_page_text.split("\n"):
    if "Summe ohne Mwst" in line:
        # Beispielzeile: "Summe ohne Mwst: CHF 123,45"
        match = re.search(r"Summe ohne Mwst.*?CHF\s*([\d.,]+)", line)
        if match:
            summe_ohne_mwst = match.group(1).strip()
        break

# --------------------------------------------------
# 1.2 Extraktion "Dieselzuschlag Aktuell" aus der letzten PDF-Seite (wird nun separat extrahiert, aber nicht im Finalcheck verwendet):
dieselzuschlag_aktuell = ""
for line in last_page_text.split("\n"):
    if "Dieselzuschlag Aktuell" in line:
        # Beispielzeile: "Dieselzuschlag Aktuell: CHF 45,67"
        match = re.search(r"Dieselzuschlag Aktuell.*?CHF\s*([\d.,]+)", line)
        if match:
            dieselzuschlag_aktuell = match.group(1).strip()
        break

# --------------------------------------------------
# 2. Überprüfung der Anzahl Fahrzeuge:
try:
    num_fahrz_file1 = int(general_info["Anzahl Fahrz."].replace(" ", ""))
except ValueError:
    num_fahrz_file1 = 0

num_cars_file2 = len(cars_data)  # Anzahl der Fahrzeuge in file2

if num_fahrz_file1 == num_cars_file2:
    anzahl_check = "OK"
else:
    anzahl_check = "NOK"

# --------------------------------------------------
# 3. Vergleich des Nettobetrags:
try:
    nettobetrag_file1 = float(general_info["Nettobetrag"].replace(",", ".").replace(" ", ""))
except ValueError:
    nettobetrag_file1 = 0.0

# Aggregierte Summe aus mehreren Spalten in file2 berechnen:
aggregate_sum_file2 = 0.0
columns_to_sum = [
    "Total", 
    "LEERFAHRT", 
    "Car Auktion Protokoll", 
    "Terminverein. Absender CarAukt", 
    "Seilwinde", 
    "Terminzuschlag", 
    "E-Fahrzeug"
]
for col in columns_to_sum:
    for value in df2[col]:
        try:
            aggregate_sum_file2 += float(str(value).replace(",", ".").replace(" ", ""))
        except (ValueError, AttributeError):
            continue

# Nettobetrag aus file1 mit der aggregierten Summe vergleichen (Toleranz 0.01):
if abs(nettobetrag_file1 - aggregate_sum_file2) < 0.01:
    nettobetrag_check = "OK"
else:
    nettobetrag_check = "NOK"

# --------------------------------------------------
# 4. Finaler Nettobetragcheck:
# Hier vergleichen wir den aus der PDF extrahierten Wert "Summe ohne MwSt" (ohne Dieselzuschlag) 
# mit der aggregierten Summe aus file2 (aggregate_sum_file2).
try:
    summe_pdf = convert_german_number(summe_ohne_mwst)
except ValueError:
    summe_pdf = 0.0

# Beide Werte auf 2 Dezimalstellen runden:
summe_pdf = round(summe_pdf, 2)
aggregate_sum_file2_rounded = round(aggregate_sum_file2, 2)

if abs(summe_pdf - aggregate_sum_file2_rounded) < 0.01:
    final_nettobetrag_check = "OK"
else:
    final_nettobetrag_check = "NOK"

# --------------------------------------------------
# 5. DataFrame für file3.xlsx erstellen mit der zusätzlichen Spalte "Final Nettobetragcheck"
df3 = pd.DataFrame({
    "Anzahl Fahrz. Check": [anzahl_check],
    "Nettobetrag Check": [nettobetrag_check],
    "Summe Total kalkuliert": [aggregate_sum_file2],
    "Summe ohne MwSt PDF": [summe_ohne_mwst],
    "Dieselzuschlag Aktuell PDF": [dieselzuschlag_aktuell],
    "Final Nettobetragcheck": [final_nettobetrag_check]
})

# file3.xlsx speichern
with pd.ExcelWriter("file3.xlsx", engine="openpyxl") as writer:
    df3.to_excel(writer, index=False)
    worksheet = writer.sheets["Sheet1"]
    uniform_width = 25
    for col_num in range(1, len(df3.columns) + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = uniform_width

# --------------------------------------------------
# Neuer Block: Vergleich von file2 mit den Daten aus ca3.xlsx und rrm.xlsx 
# und Ausgabe der Ergebnisse in file4.xlsx
# --------------------------------------------------

import pandas as pd
from openpyxl.utils import get_column_letter

# 1. Laden der Referenzdaten aus den Excel-Dateien ca3.xlsx und rrm.xlsx
df_ca3 = pd.read_excel("ca3.xlsx")
df_rrm = pd.read_excel("rrm.xlsx")

# 2. Hilfsfunktion zur Konvertierung von Zahlen (Umwandeln vom deutschen Format in Float)
def convert_numeric(value):
    """
    Konvertiert einen numerischen Wert (als String oder Zahl) vom deutschen Format
    in einen Float.
      - Fehlt der Wert (NaN oder leer), wird None zurückgegeben.
      - Entfernt unsichtbare Leerzeichen (z. B. non-breaking spaces).
      - Liegt nur ein Komma vor (z. B. "1,5"), wird das Komma durch einen Punkt ersetzt.
      - Sind Punkt und Komma vorhanden (Punkt als Tausendertrenner), werden zuerst alle Punkte entfernt,
        dann das Komma durch einen Punkt ersetzt.
    """
    try:
        if pd.isna(value):
            return None
        s = str(value).replace("\xa0", "").strip()
        if s == "":
            return None
        if ',' in s and '.' not in s:
            s = s.replace(',', '.')
        elif '.' in s and ',' in s:
            s = s.replace('.', '')
            s = s.replace(',', '.')
        return round(float(s), 2)
    except Exception:
        return None

def compare_numeric_values(val1, val2):
    """
    Vergleicht zwei numerische Werte, nachdem sie umgewandelt wurden.
    Sind beide Werte leer (None), wird "OK" zurückgegeben.
    Ist nur einer leer, liefert die Funktion "NOK".
    Ansonsten werden die Werte (auf 2 Dezimalstellen gerundet) verglichen.
    """
    n1 = convert_numeric(val1)
    n2 = convert_numeric(val2)
    if n1 is None and n2 is None:
        return "OK"
    if n1 is None or n2 is None:
        return "NOK"
    return "OK" if n1 == n2 else "NOK"

def compare_text_values(v1, v2):
    """
    Vergleicht zwei Textwerte.
    Fehlt ein Wert (pd.isna), wird er als leerer String berücksichtigt.
    Zusätzlich werden unsichtbare Leerzeichen entfernt.
    Beide Werte werden getrimmt und in Kleinbuchstaben umgewandelt.
    """
    s1 = "" if pd.isna(v1) else str(v1).replace("\xa0", "").strip().lower()
    s2 = "" if pd.isna(v2) else str(v2).replace("\xa0", "").strip().lower()
    return "OK" if s1 == s2 else "NOK"

def longest_common_substring(s1, s2):
    m, n = len(s1), len(s2)
    dp = [[0]*(n+1) for _ in range(m+1)]
    longest = 0
    for i in range(1, m+1):
        for j in range(1, n+1):
            if s1[i-1] == s2[j-1]:
                dp[i][j] = dp[i-1][j-1] + 1
                if dp[i][j] > longest:
                    longest = dp[i][j]
            else:
                dp[i][j] = 0
    return longest

def compare_city(val1, val2):
    # Konvertiere die Eingabewerte in Kleinbuchstaben,
    # entferne unsichtbare Leerzeichen (z. B. non-breaking spaces) und trimme den Whitespace.
    s1 = "" if pd.isna(val1) else str(val1).replace("\xa0", "").strip().lower()
    s2 = "" if pd.isna(val2) else str(val2).replace("\xa0", "").strip().lower()
    
    # Falls beide Werte leer sind, werden sie als gleich angesehen.
    if s1 == "" and s2 == "":
        return "OK"
    
    # Sonderregel: Wenn ein Wert "nebikon" und der andere "altishofen" (oder umgekehrt) ist,
    # werden diese als gleich betrachtet.
    if (s1 == "nebikon" and s2 == "altishofen") or (s1 == "altishofen" and s2 == "nebikon"):
        return "OK"
    
    # Wenn die Länge der längsten gemeinsamen Teilzeichenkette mindestens 5 Zeichen beträgt,
    # werden die Werte als ähnlich betrachtet.
    if longest_common_substring(s1, s2) >= 4:
        return "OK"
    
    # Ansonsten gelten die Werte als unterschiedlich.
    return "NOK"

def compare_loadingcity(val1, val2):
    """
    Vergleicht zwei Werte aus dem Feld Loadingcity.
    Entfernt unsichtbare Leerzeichen, trimmt und wandelt in Kleinbuchstaben um.
    Falls ein Wert "nebikon" und der andere "altishofen" (bzw. umgekehrt) ist, wird "OK" zurückgegeben.
    Andernfalls erfolgt der normale Textvergleich.
    """
    s1 = "" if pd.isna(val1) else str(val1).replace("\xa0", "").strip().lower()
    s2 = "" if pd.isna(val2) else str(val2).replace("\xa0", "").strip().lower()
    if (s1 == "nebikon" and s2 == "altishofen") or (s1 == "altishofen" and s2 == "nebikon"):
        return "OK"
    return "OK" if s1 == s2 else "NOK"

# 3. Liste zur Speicherung der Vergleichsergebnisse erstellen
compare_results = []

# 4. Iteration über alle Zeilen in file2 (df2)
for idx, row in df2.iterrows():
    invoice_nr = str(row["InvoiceNr"]).strip()
    
    # Auswahl der Referenzdatei anhand der Länge der InvoiceNr:
    # - 5-stellige InvoiceNr → rrm.xlsx
    # - 6-stellige InvoiceNr → ca3.xlsx
    if len(invoice_nr) == 5:
        df_reference = df_rrm
    elif len(invoice_nr) == 6:
        df_reference = df_ca3
    else:
        df_reference = None

    # Suche der passenden Zeile in der Referenzdatei anhand der Spalte "invoice"
    if df_reference is not None:
        matching_rows = df_reference[df_reference["invoice"].astype(str).str.strip() == invoice_nr]
        if len(matching_rows) > 0:
            ref = matching_rows.iloc[0]
        else:
            ref = None
    else:
        ref = None

    # Falls eine Referenzzeile gefunden wurde, erfolgt der Vergleich.
    if ref is not None:
        auftraggeber_vergleich    = compare_text_values(row["Auftraggeber"], ref["Auftraggeber"])
        vin_vergleich             = compare_text_values(row["VIN"], ref["vin"])
        loadingcity_vergleich = compare_city(row["Loadingcity"], ref["loadingcity"])
        delivercity_vergleich = compare_city(row["Delivercity"], ref["delivercity"])
        faktor_vergleich          = compare_numeric_values(row["Faktor"], ref["Faktor"])
        transportrpeis_vergleich  = compare_numeric_values(row["Total"], ref["Gallikerpreis"])
        telavis_vergleich         = compare_numeric_values(row["Terminverein. Absender CarAukt"], ref["Telavis"])
        seilwinde_vergleich       = compare_numeric_values(row["Seilwinde"], ref["Seilwinde"])
        terminzuschlag_vergleich  = compare_numeric_values(row["Terminzuschlag"], ref["Terminzuschlag"])
        # Änderung: Für E-FahrzeugVergleich wird nun Car Auktion Protokoll aus file2, 
        # verglichen mit EÜbername aus der Referenzdatei.
        efahrzeug_vergleich       = compare_numeric_values(row["Car Auktion Protokoll"], ref["EÜbernahme"])
        
        leerfahrt_vergleich       = compare_text_values(row["LEERFAHRT"], ref["Leerfahrt"])
    else:
        auftraggeber_vergleich = vin_vergleich = loadingcity_vergleich = delivercity_vergleich = "NOK"
        faktor_vergleich = transportrpeis_vergleich = telavis_vergleich = seilwinde_vergleich = terminzuschlag_vergleich = "NOK"
        efahrzeug_vergleich = leerfahrt_vergleich = "NOK"
    
    # 5. Generierung der Bemerkungen-Spalte mit detaillierter Vergleichslogik
    bemerkungen = []

    # Auftraggeber und VIN Vergleich
    if auftraggeber_vergleich == "NOK":
        bemerkungen.append("Unterschiedliche Auftraggeber")
    if vin_vergleich == "NOK":
        bemerkungen.append("VIN stimmt nicht überein")

    # Lade- und Lieferorte Vergleich
    if loadingcity_vergleich == "NOK":
        bemerkungen.append("Unterschiedliche Ladeorte")
    if delivercity_vergleich == "NOK":
        bemerkungen.append("Unterschiedliche Lieferorte")

    # Faktor Vergleich mit spezifischer Logik
    faktor_file2 = convert_numeric(row["Faktor"])
    try:
        faktor_ref = convert_numeric(ref["Faktor"])
    except (TypeError, KeyError):
        faktor_ref = None

    if faktor_vergleich == "NOK":
        if faktor_file2 is None and faktor_ref is not None:
            bemerkungen.append("Faktor ist vorhanden nur auf CA3 (RRM)")
        elif faktor_file2 is not None and faktor_ref is None:
            bemerkungen.append("Faktor ist vorhanden nur in der Gallikerrechnung")
        elif faktor_file2 < faktor_ref:
            bemerkungen.append("Faktor in Galliker Rechnung ist geringer als auf CA3 (RRM)")
        elif faktor_file2 > faktor_ref:
            bemerkungen.append("Faktor in Galliker Rechnung ist größer als auf CA3 (RRM)")

    # Transportpreis Vergleich mit spezifischer Logik
    transportpreis_file2 = convert_numeric(row["Total"])
    try:
        transportpreis_ref = convert_numeric(ref["Gallikerpreis"])
    except (TypeError, KeyError):
        transportpreis_ref = None

    if transportrpeis_vergleich == "NOK":
        if transportpreis_file2 is None and transportpreis_ref is not None:
            bemerkungen.append("Transportpreis ist vorhanden nur auf CA3 (RRM)")
        elif transportpreis_file2 is not None and transportpreis_ref is None:
            bemerkungen.append("Transportpreis ist vorhanden nur in der Gallikerrechnung")
        elif transportpreis_file2 < transportpreis_ref:
            bemerkungen.append("Transportpreis in Galliker Rechnung ist geringer als kalkulatorischer Transportpreis auf CA3 (RRM)")
        elif transportpreis_file2 > transportpreis_ref:
            bemerkungen.append("Transportpreis in Galliker Rechnung ist größer als kalkulatorischer Transportpreis auf CA3 (RRM)")

    # Telavis Vergleich mit spezifischer Logik
    if telavis_vergleich == "NOK":
        telavis_file2 = "" if pd.isna(row["Terminverein. Absender CarAukt"]) else str(row["Terminverein. Absender CarAukt"]).strip()
        try:
            telavis_ref_val = "" if pd.isna(ref["Telavis"]) else str(ref["Telavis"]).strip()
        except (TypeError, KeyError):
            telavis_ref_val = ""
        if telavis_file2 == "" and telavis_ref_val != "":
            bemerkungen.append("Telavis ist vorhanden nur auf CA3 (RRM)")
        elif telavis_file2 != "" and telavis_ref_val == "":
            bemerkungen.append("Telavis ist vorhanden nur in der Gallikerrechnung")
        else:
            bemerkungen.append("Terminvereinbarung weicht ab")

    # Seilwinde Vergleich
    if seilwinde_vergleich == "NOK":
        seilwinde_file2 = "" if pd.isna(row["Seilwinde"]) else str(row["Seilwinde"]).strip()
        try:
            seilwinde_ref_val = "" if pd.isna(ref["Seilwinde"]) else str(ref["Seilwinde"]).strip()
        except (TypeError, KeyError):
            seilwinde_ref_val = ""
        if seilwinde_file2 == "" and seilwinde_ref_val != "":
            bemerkungen.append("Seilwinde ist vorhanden nur auf CA3 (RRM)")
        elif seilwinde_file2 != "" and seilwinde_ref_val == "":
            bemerkungen.append("Seilwinde ist vorhanden nur in der Gallikerrechnung")
        else:
            bemerkungen.append("Seilwinde unterschiedlich")

    # Terminzuschlag Vergleich
    if terminzuschlag_vergleich == "NOK":
        terminzuschlag_file2 = "" if pd.isna(row["Terminzuschlag"]) else str(row["Terminzuschlag"]).strip()
        try:
            terminzuschlag_ref_val = "" if pd.isna(ref["Terminzuschlag"]) else str(ref["Terminzuschlag"]).strip()
        except (TypeError, KeyError):
            terminzuschlag_ref_val = ""
        if terminzuschlag_file2 == "" and terminzuschlag_ref_val != "":
            bemerkungen.append("Terminzuschlag ist vorhanden nur auf CA3 (RRM)")
        elif terminzuschlag_file2 != "" and terminzuschlag_ref_val == "":
            bemerkungen.append("Terminzuschlag ist vorhanden nur in der Gallikerrechnung")
        else:
            bemerkungen.append("Terminzuschlag weicht ab")

    # E-Fahrzeug Vergleich
    if efahrzeug_vergleich == "NOK":
        efahrzeug_file2 = "" if pd.isna(row["Car Auktion Protokoll"]) else str(row["Car Auktion Protokoll"]).strip()
        try:
            efahrzeug_ref_val = "" if pd.isna(ref["EÜbernahme"]) else str(ref["EÜbernahme"]).strip()
        except (TypeError, KeyError):
            efahrzeug_ref_val = ""
        if efahrzeug_file2 == "" and efahrzeug_ref_val != "":
            bemerkungen.append("E-Fahrzeug ist vorhanden nur auf CA3 (RRM)")
        elif efahrzeug_file2 != "" and efahrzeug_ref_val == "":
            bemerkungen.append("E-Fahrzeug ist vorhanden nur in der Gallikerrechnung")
        else:
            bemerkungen.append("E-Fahrzeug unterschiedlich")

    # Leerfahrt Vergleich
    if leerfahrt_vergleich == "NOK":
        leerfahrt_file2 = "" if pd.isna(row["LEERFAHRT"]) else str(row["LEERFAHRT"]).strip()
        try:
            leerfahrt_ref_val = "" if pd.isna(ref["Leerfahrt"]) else str(ref["Leerfahrt"]).strip()
        except (TypeError, KeyError):
            leerfahrt_ref_val = ""
        if leerfahrt_file2 == "" and leerfahrt_ref_val != "":
            bemerkungen.append("Leerfahrt ist vorhanden nur auf CA3 (RRM)")
        elif leerfahrt_file2 != "" and leerfahrt_ref_val == "":
            bemerkungen.append("Leerfahrt ist vorhanden nur in der Gallikerrechnung")
        else:
            bemerkungen.append("Leerfahrt bitte prüfen")

    # Die Bemerkungen werden in eine Zeichenkette umgewandelt, getrennt durch Kommas
    bemerkungen_text = ", ".join(bemerkungen) if bemerkungen else ""

    # 6. Zusammenstellung der neuen Zeile mit den Originaldaten aus file2 und den Vergleichsergebnissen
    new_row = {
        "InvoiceNr": row["InvoiceNr"],
        "Auftraggeber": row["Auftraggeber"],  # Auftraggeber aus file2
        "VIN": row["VIN"],
        "Model": row["Model"],
        "Faktor": row["Faktor"],
        "Total": row["Total"],
        "Loadingcity": row["Loadingcity"],
        "Delivercity": row["Delivercity"],
        "LEERFAHRT": row["LEERFAHRT"],
        "Car Auktion Protokoll": row["Car Auktion Protokoll"],
        "Terminverein. Absender CarAukt": row["Terminverein. Absender CarAukt"],
        "Seilwinde": row["Seilwinde"],
        "Terminzuschlag": row["Terminzuschlag"],
        "E-Fahrzeug": row["E-Fahrzeug"],
        # Vergleichsergebnisse
        "AuftraggeberVergleich": auftraggeber_vergleich,
        "VINVergleich": vin_vergleich,
        "LoadingcityVergleich": loadingcity_vergleich,
        "DelivercityVergleich": delivercity_vergleich,
        "FaktorVergleich": faktor_vergleich,
        "TransportrpeisVergleich": transportrpeis_vergleich,
        "TelavisVergleich": telavis_vergleich,
        "SeilwindeVergleich": seilwinde_vergleich,
        "TerminzuschlagVergleich": terminzuschlag_vergleich,
        "E-FahrzeugVergleich": efahrzeug_vergleich,
        "LeerfahrtVergleich": leerfahrt_vergleich,
        "Bemerkungen": bemerkungen_text
    }
    compare_results.append(new_row)

# 7. Erstellen eines neuen DataFrames für file4 mit den Vergleichsergebnissen
df4 = pd.DataFrame(compare_results)

# 8. Speichern des Ergebnisses in file4.xlsx mit einheitlicher Spaltenbreite
with pd.ExcelWriter("file4.xlsx", engine="openpyxl") as writer:
    df4.to_excel(writer, index=False)
    worksheet = writer.sheets["Sheet1"]
    for col_num, column_name in enumerate(df4.columns, start=1):
        if column_name == "Bemerkungen":
            worksheet.column_dimensions[get_column_letter(col_num)].width = 75  # x3 Breite
        else:
            worksheet.column_dimensions[get_column_letter(col_num)].width = 25  # normale Breite

print("✅ Die Info ist in file1.xlsx, file2.xlsx und file3.xlsx erfolgreich gespeichert!")
print("✅ Der Vergleich wurde in file4.xlsx erfolgreich gespeichert!")